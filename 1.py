# -*- coding:utf-8 -*-
import sys
import os
import numpy as np
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QFormLayout,
                             QLineEdit, QLabel, QSpinBox, QDoubleSpinBox, QPushButton,
                             QTableWidget, QTableWidgetItem, QFileDialog, QTextEdit, QGroupBox,
                             QHBoxLayout, QVBoxLayout, QComboBox, QInputDialog, QMessageBox,
                             QScrollArea, QSplitter, QMenu, QCheckBox, QSizePolicy)
from PyQt5.QtCore import Qt, QRect, QTimer, QSize
from PyQt5.QtGui import QPixmap, QFont, QColor, QPen, QBrush, QImage, QIcon, QPainter
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import openpyxl
import matplotlib

matplotlib.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei"]
matplotlib.rcParams["axes.unicode_minus"] = False


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


COLOR_BLUE_BORDER = "#A4D0FF"
COLOR_YELLOW_BORDER = "#FFE082"
COLOR_PURPLE_BORDER = "#CE93D8"
COLOR_GREEN_BORDER = "#A5D6A7"
COLOR_ORANGE_BORDER = "#FFCC80"
COLOR_BOTTOM_BORDER = "#B39DDB"
GROUP_BG = "#F5F9FF"
GROUP_BORDER = "#E0FFF"

#


class CollapsiblePanel(QWidget):
    def __init__(self, title, panel_id, parent=None):
        super().__init__(parent)
        self.title = title
        self.panel_id = panel_id
        self.is_expanded = False
        self.parent_window = parent

        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)

        self.btn_title = QPushButton(title)
        self.btn_title.setStyleSheet("""
            QPushButton {
                background-color: #EBF4FF;
                border: 1px solid #D0E7FF;
                border-radius: 6px;
                padding: 10px 15px;
                text-align: left;
                font-weight: bold;
                font-size: 15px;
                color: #2176D3;
            }
            QPushButton:hover {
                background-color: #E0EFFF;
            }
        """)
        self.btn_title.setFixedHeight(45)
        self.btn_title.clicked.connect(self.toggle)
        self.main_layout.addWidget(self.btn_title)

        self.content_area = QWidget()
        self.content_area.setStyleSheet(
            f"background-color: {GROUP_BG}; border: 1px solid {GROUP_BORDER}; border-top: none; border-bottom-left-radius: 6px; border-bottom-right-radius: 6px;")
        self.content_layout = QVBoxLayout(self.content_area)
        self.content_layout.setContentsMargins(18, 20, 18, 20)
        self.content_layout.setSpacing(15)
        self.content_area.hide()
        self.main_layout.addWidget(self.content_area)

        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)

    def toggle(self):
        self.is_expanded = not self.is_expanded
        self.content_area.setVisible(self.is_expanded)

        if self.is_expanded:
            self.btn_title.setStyleSheet("""
                QPushButton {
                    background-color: #F1F8E9;
                    border: 1px solid #81C784;
                    border-radius: 6px;
                    border-bottom-left-radius: 0px;
                    border-bottom-right-radius: 0px;
                    padding: 10px 15px;
                    text-align: left;
                    font-weight: bold;
                    font-size: 15px;
                    color: #388E3C;
                }
            """)
            if self.parent_window:
                self.parent_window.on_panel_expanded(self.panel_id)
        else:
            self.btn_title.setStyleSheet("""
                QPushButton {
                    background-color: #EBF4FF;
                    border: 1px solid #D0E7FF;
                    border-radius: 6px;
                    padding: 10px 15px;
                    text-align: left;
                    font-weight: bold;
                    font-size: 15px;
                    color: #2176D3;
                }
                QPushButton:hover {
                    background-color: #E0EFFF;
                }
            """)
            if self.parent_window:
                self.parent_window.on_panel_collapsed(self.panel_id)

        self.updateGeometry()
        if self.parent_window:
            self.parent_window.scroll_content.updateGeometry()

    def collapse(self):
        if self.is_expanded:
            self.is_expanded = False
            self.content_area.hide()
            self.btn_title.setStyleSheet("""
                QPushButton {
                    background-color: #EBF4FF;
                    border: 1px solid #D0E7FF;
                    border-radius: 6px;
                    padding: 10px 15px;
                    text-align: left;
                    font-weight: bold;
                    font-size: 15px;
                    color: #2176D3;
                }
                QPushButton:hover {
                    background-color: #E0EFFF;
                }
            """)
            if self.parent_window:
                self.parent_window.on_panel_collapsed(self.panel_id)
            self.updateGeometry()

    def add_widget(self, widget):
        self.content_layout.addWidget(widget)

    def add_layout(self, layout):
        self.content_layout.addLayout(layout)


class ReportProgressBar(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.course_name = ""
        self.percent = 0.0
        self.setFixedHeight(110)

    def set_data(self, course_name, percent):
        self.course_name = course_name
        self.percent = max(0.0, min(100.0, percent))
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        w = self.width()
        top_text_h = 26
        bar_h = 36
        radius = bar_h // 2

        painter.setPen(QColor("#2176D3"))
        font = QFont()
        font.setPointSize(11)
        font.setBold(True)
        painter.setFont(font)
        show_name = self.course_name if self.course_name else "未填写课程名称"
        painter.drawText(QRect(0, 0, w, top_text_h), Qt.AlignCenter,
                         f"课程名称:{show_name}  {self.percent:.1f}%")

        bar_rect = QRect(3, top_text_h, w - 6, bar_h)
        painter.setPen(Qt.NoPen)
        painter.setBrush(QColor("#ffffff"))
        painter.drawRoundedRect(bar_rect, radius, radius)

        fill_width = int((w - 6) * self.percent / 100)
        if fill_width > 0:
            fill_rect = QRect(3, top_text_h, fill_width, bar_h)
            painter.setBrush(QColor("#A4D0FF"))
            painter.drawRoundedRect(fill_rect, radius, radius)

        painter.setPen(QPen(QColor("#A4D0FF"), 2))
        painter.setBrush(Qt.NoBrush)
        painter.drawRoundedRect(bar_rect, radius, radius)


class MplCanvasRadar(FigureCanvas):
    def __init__(self, parent=None, width=3.3, height=2.4, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = self.fig.add_subplot(111, polar=True)
        super().__init__(self.fig)


class MplCanvasPie(FigureCanvas):
    def __init__(self, parent=None, width=3.3, height=2.4, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = self.fig.add_subplot(111)
        super().__init__(self.fig)


class MplCanvasBar(FigureCanvas):
    def __init__(self, parent=None, width=3.3, height=2.4, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = self.fig.add_subplot(111)
        super().__init__(self.fig)


class NoWheelSpin(QSpinBox):
    def __init__(self):
        super().__init__()
        self.setButtonSymbols(QSpinBox.NoButtons)
        self.setStyleSheet(
            "QSpinBox{padding:6px 7px;border:1px solid #D0E7FF;border-radius:4px;background-color: white;}")
        self.setFocusPolicy(Qt.StrongFocus)

    def wheelEvent(self, event):
        event.ignore()


class NoWheelDoubleSpin(QDoubleSpinBox):
    def __init__(self):
        super().__init__()
        self.setButtonSymbols(QDoubleSpinBox.NoButtons)
        self.setStyleSheet(
            "QDoubleSpinBox{padding:6px 7px;border:1px solid #D0E7FF;border-radius:4px;background-color: white;}")
        self.setFocusPolicy(Qt.StrongFocus)

    def wheelEvent(self, event):
        event.ignore()


class NoWheelLineEdit(QLineEdit):
    def __init__(self):
        super().__init__()
        self.setStyleSheet(
            "QLineEdit{padding:6px 7px;border:1px solid #D0E7FF;border-radius:4px;background-color: white;}")
        self.setFocusPolicy(Qt.StrongFocus)
        self.setReadOnly(False)

    def wheelEvent(self, event):
        event.ignore()


class NoWheelTextEdit(QTextEdit):
    def __init__(self):
        super().__init__()
        self.setStyleSheet(
            "QTextEdit{padding:6px 7px;border:1px solid #D0E7FF;border-radius:4px;background-color: white;}")
        self.setFocusPolicy(Qt.StrongFocus)
        self.setReadOnly(False)

    def wheelEvent(self, event):
        event.ignore()


class NoWheelComboBox(QComboBox):
    def __init__(self):
        super().__init__()
        self.setStyleSheet(
            "QComboBox{padding:6px 7px;border:1px solid #D0E7FF;border-radius:4px;background-color: white;}")
        self.setFocusPolicy(Qt.StrongFocus)

    def wheelEvent(self, event):
        event.ignore()


class KnowledgeCheckBox(QCheckBox):
    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.setFocusPolicy(Qt.StrongFocus)


class MainWin(QMainWindow):
    def __init__(self):
        super().__init__()
        global_font = QFont("Microsoft YaHei", 10)
        app = QApplication.instance()
        app.setFont(global_font)
        self.setWindowTitle("测试分析报告编辑器")
        self.move(80, 40)
        self.setMinimumSize(1200, 800)
        self.resize(1920, 1380)
        self.loss_lib = ["思路错误", "细节出错", "计算错误", "概念错误"]
        self.loss_data = {}
        self.prog_item_list = []
        self.prog_full_sum = 0.0
        self.prog_get_sum = 0.0

        # 恢复自动刷新定时器，确保实时更新
        self.timer_refresh = QTimer()
        self.timer_refresh.setInterval(500)  # 缩短刷新间隔到500ms，更及时
        self.timer_refresh.timeout.connect(self.sync_preview)
        self.timer_refresh.start()

        self.cache_bar_text = ""
        self.know_check_list = []
        self.all_panels = {}
        self.open_panels_history = []
        self.max_open_panels = 3

        self.init_ui()
        self.all_panels["stu"].toggle()
        self.all_panels["score"].toggle()
        self.all_panels["prog"].toggle()

    def on_panel_expanded(self, panel_id):
        if panel_id in self.open_panels_history:
            self.open_panels_history.remove(panel_id)
        self.open_panels_history.append(panel_id)
        while len(self.open_panels_history) > self.max_open_panels:
            oldest_panel_id = self.open_panels_history.pop(0)
            self.all_panels[oldest_panel_id].collapse()

    def on_panel_collapsed(self, panel_id):
        if panel_id in self.open_panels_history:
            self.open_panels_history.remove(panel_id)

    def auto_open_panel(self, panel_id):
        if panel_id in self.all_panels and not self.all_panels[panel_id].is_expanded:
            self.all_panels[panel_id].toggle()

    def refresh_knowledge_height(self):
        count = len(self.know_check_list)
        base_height = 60
        per_height = 26
        new_height = base_height + count * per_height
        self.container_check.setMinimumHeight(new_height)
        self.container_check.updateGeometry()

    def get_radar_evaluation(self):
        labels = ["概念理解力", "计算能力", "模型应用", "审题能力", "知识体系"]
        scores = [self.radar_dict[name].value() for name in labels]
        avg = sum(scores) / len(scores)
        if avg >= 90:
            star = 5;
            title = "顶尖优秀";
            comment = "五维能力均衡拔尖，基础扎实，解题思路清晰，审题严谨，计算稳定，知识体系完整，综合能力极强。"
        elif avg >= 80:
            star = 4;
            title = "良好偏优";
            comment = "整体能力良好，基础牢固，个别维度略有短板，常规题型稳定发挥，变式题具备独立思考能力。"
        elif avg >= 70:
            star = 3;
            title = "中等达标";
            comment = "基础能力合格，基础题得分稳定，中档题偶有卡顿，针对性补强薄弱项即可稳步提升。"
        elif avg >= 60:
            star = 2;
            title = "临界及格";
            comment = "仅掌握基础内容，中档题与拓展题完成困难，多项能力存在漏洞，需从基础查漏补缺。"
        else:
            star = 0;
            title = "待夯实基础";
            comment = "整体基础薄弱，多维度存在知识断层，优先夯实课本基础概念、计算与例题。"
        return star, title, comment

    def generate_analysis(self, data_dict):
        core = ""
        improve = ""
        sort_list = sorted(data_dict.items(), key=lambda x: -x[1])
        for typ, val in sort_list:
            if val <= 0: continue
            if typ == "思路错误":
                core += "• 解题逻辑不清晰，题型拆解能力弱，无法建立正确解题框架\n"
                improve += "• 梳理题型模板，强化分步拆解，练习逆向推导与验算\n"
            elif typ == "细节出错":
                core += "• 审题粗心、书写不规范、边界条件遗漏，低级失误偏多\n"
                improve += "• 圈画关键词，规范书写步骤，逐行核对，养成自查习惯\n"
            elif typ == "计算错误":
                core += "• 运算准确率不足，心算跳步多，符号与优先级处理混乱\n"
                improve += "• 强化草稿规范，分步计算，做完强制验算复核\n"
            elif typ == "概念错误":
                core += "• 基础概念理解模糊，知识点混淆，公式与定义应用错误\n"
                improve += "• 回归课本夯实概念，区分易混点，建立知识体系\n"
            else:
                core += f"• {typ}失分占比偏高，需针对性专项训练\n"
                improve += f"• 针对{typ}整理错题，定点巩固练习\n"
        if not core:
            core = "暂无失分数据"
            improve = "正常发挥，继续保持"
        return core.strip(), improve.strip()

    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        splitter_main = QSplitter(Qt.Horizontal)
        splitter_main.setStyleSheet("QSplitter::handle{background:#cccccc;width:7px;}")
        splitter_main.setChildrenCollapsible(False)

        left_container = QWidget()
        left_container.setStyleSheet("background-color:#ffffff")
        left_layout = QVBoxLayout(left_container)
        left_layout.setContentsMargins(0, 0, 0, 0)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setStyleSheet("border:none;")
        self.scroll_content = QWidget()
        self.scroll_content.setStyleSheet("background:#ffffff;")
        self.scroll_v = QVBoxLayout(self.scroll_content)
        self.scroll_v.setSpacing(10)
        self.scroll_v.setContentsMargins(14, 14, 14, 14)

        panel_stu = CollapsiblePanel("【学生信息】", "stu", self)
        self.all_panels["stu"] = panel_stu
        lay_stu = QFormLayout()
        lay_stu.setLabelAlignment(Qt.AlignLeft)
        lay_stu.setSpacing(15)
        lay_stu.setHorizontalSpacing(26)
        self.in_name = NoWheelLineEdit()
        self.in_course = NoWheelLineEdit()
        self.in_time = NoWheelLineEdit()
        self.in_teacher = NoWheelLineEdit()

        # 学生信息实时更新绑定（双重保险：信号+定时器）
        self.in_name.textChanged.connect(self.sync_preview)
        self.in_course.textChanged.connect(self.sync_preview)
        self.in_time.textChanged.connect(self.sync_preview)
        self.in_teacher.textChanged.connect(self.sync_preview)

        lay_stu.addRow(QLabel("学生姓名："), self.in_name)
        lay_stu.addRow(QLabel("课程名称："), self.in_course)
        lay_stu.addRow(QLabel("测试时间："), self.in_time)
        lay_stu.addRow(QLabel("教师姓名："), self.in_teacher)
        panel_stu.add_layout(lay_stu)
        self.scroll_v.addWidget(panel_stu)

        panel_score = CollapsiblePanel("【得分统计】", "score", self)
        self.all_panels["score"] = panel_score
        lay_score = QFormLayout()
        lay_score.setLabelAlignment(Qt.AlignLeft)
        lay_score.setSpacing(15)
        lay_score.setHorizontalSpacing(26)
        self.sp_full = NoWheelDoubleSpin()
        self.sp_full.setRange(0, 999)
        self.sp_get = NoWheelDoubleSpin()
        self.sp_get.setRange(0, 999)
        self.sp_obj_r = NoWheelSpin()
        self.sp_obj_w = NoWheelSpin()
        self.sp_prog_cnt = NoWheelSpin()
        self.sp_prog_cnt.valueChanged.connect(lambda: self.auto_open_panel("prog"))
        btn_h = QHBoxLayout()
        btn_h.addStretch()
        self.btn_gen_prog = QPushButton("生成题目条目")
        self.btn_gen_prog.setStyleSheet("""
            QPushButton {padding: 6px 12px;border: 1px solid #D0E7FF;border-radius: 4px;background-color: white;}
            QPushButton:hover {background-color: #EBF4FF;}
        """)
        self.btn_gen_prog.clicked.connect(self.create_prog_items)
        btn_h.addWidget(self.btn_gen_prog)

        # 综合得分实时更新绑定（双重保险：信号+定时器）
        self.sp_full.valueChanged.connect(self.sync_preview)
        self.sp_get.valueChanged.connect(self.sync_preview)
        self.sp_obj_r.valueChanged.connect(self.sync_preview)
        self.sp_obj_w.valueChanged.connect(self.sync_preview)
        self.sp_prog_cnt.valueChanged.connect(self.sync_preview)

        lay_score.addRow(QLabel("试卷总分："), self.sp_full)
        lay_score.addRow(QLabel("实际得分："), self.sp_get)
        lay_score.addRow(QLabel("客观答对："), self.sp_obj_r)
        lay_score.addRow(QLabel("客观答错："), self.sp_obj_w)
        lay_score.addRow(QLabel("计算题："), self.sp_prog_cnt)  # 统一文字为"编程题"
        lay_score.addRow("", btn_h)
        panel_score.add_layout(lay_score)
        self.scroll_v.addWidget(panel_score)

        panel_prog = CollapsiblePanel("【计算题明细】", "prog", self)
        self.all_panels["prog"] = panel_prog
        self.lay_prog_wrap = QVBoxLayout()
        self.lay_prog_wrap.setSpacing(9)
        panel_prog.add_layout(self.lay_prog_wrap)
        self.scroll_v.addWidget(panel_prog)

        panel_radar = CollapsiblePanel("【能力雷达设置】满分100分", "radar", self)
        self.all_panels["radar"] = panel_radar
        lay_radar = QFormLayout()
        lay_radar.setLabelAlignment(Qt.AlignLeft)
        lay_radar.setSpacing(15)
        lay_radar.setHorizontalSpacing(26)
        self.radar_dict = {}
        radar_names = ["概念理解力", "计算能力", "模型应用", "审题能力", "知识体系"]
        for name in radar_names:
            sp = NoWheelDoubleSpin()
            sp.setRange(0, 100)
            sp.setSingleStep(1)
            sp.valueChanged.connect(lambda: self.auto_open_panel("bottom"))
            self.radar_dict[name] = sp
            lay_radar.addRow(QLabel(f"{name}："), sp)
        btn_radar_h = QHBoxLayout()
        btn_radar_h.addStretch()
        self.btn_draw_radar = QPushButton("刷新雷达图与自动建议")
        self.btn_draw_radar.setStyleSheet("""
            QPushButton {padding: 6px 12px;border: 1px solid #D0E7FF;border-radius: 4px;background-color: white;}
            QPushButton:hover {background-color: #EBF4FF;}
        """)
        self.btn_draw_radar.clicked.connect(self.on_refresh_radar)
        btn_radar_h.addWidget(self.btn_draw_radar)
        lay_radar.addRow("", btn_radar_h)
        panel_radar.add_layout(lay_radar)
        self.scroll_v.addWidget(panel_radar)

        panel_know = CollapsiblePanel("【知识点掌握统计】", "know", self)
        self.all_panels["know"] = panel_know
        lay_know = QFormLayout()
        lay_know.setLabelAlignment(Qt.AlignLeft)
        lay_know.setSpacing(15)
        lay_know.setHorizontalSpacing(26)
        self.sp_k_all = NoWheelSpin()
        self.sp_k_good = NoWheelSpin()
        self.sp_k_mid = NoWheelSpin()
        self.sp_k_bad = NoWheelSpin()
        lay_know.addRow(QLabel("知识点总数："), self.sp_k_all)
        lay_know.addRow(QLabel("掌握较好(L5)："), self.sp_k_good)
        lay_know.addRow(QLabel("中等掌握(L2-L4)："), self.sp_k_mid)
        lay_know.addRow(QLabel("掌握较差(L0)："), self.sp_k_bad)
        btn_know_h = QHBoxLayout()
        btn_know_h.addStretch()
        self.btn_draw_pie = QPushButton("刷新环形分布图")
        self.btn_draw_pie.setStyleSheet("""
            QPushButton {padding: 6px 12px;border: 1px solid #D0E7FF;border-radius: 4px;background-color: white;}
            QPushButton:hover {background-color: #EBF4FF;}
        """)
        self.btn_draw_pie.clicked.connect(self.sync_preview)
        lay_know.addRow("", btn_know_h)
        panel_know.add_layout(lay_know)
        self.scroll_v.addWidget(panel_know)

        panel_loss = CollapsiblePanel("【失分录入】", "loss", self)
        self.all_panels["loss"] = panel_loss
        lay_loss_top = QHBoxLayout()
        lay_loss_top.setSpacing(8)
        self.com_loss = NoWheelComboBox()
        self.com_loss.addItems(self.loss_lib)
        self.sp_loss_per = NoWheelDoubleSpin()
        self.sp_loss_per.setRange(0, 100)
        self.btn_add_loss = QPushButton("添加")
        self.btn_add_loss.setStyleSheet("""
            QPushButton {padding: 6px 12px;border: 1px solid #D0E7FF;border-radius: 4px;background-color: white;}
            QPushButton:hover {background-color: #EBF4FF;}
        """)
        self.btn_custom_loss = QPushButton("自定义原因")
        self.btn_custom_loss.setStyleSheet("""
            QPushButton {padding: 6px 12px;border: 1px solid #D0E7FF;border-radius: 4px;background-color: white;}
            QPushButton:hover {background-color: #EBF4FF;}
        """)
        lay_loss_top.addWidget(self.com_loss)
        lay_loss_top.addWidget(QLabel("占比%"))
        lay_loss_top.addWidget(self.sp_loss_per)
        lay_loss_top.addWidget(self.btn_add_loss)
        lay_loss_top.addWidget(self.btn_custom_loss)

        self.table_loss = QTableWidget()
        self.table_loss.setColumnCount(2)
        self.table_loss.setHorizontalHeaderLabels(["失分原因", "占比%"])
        self.table_loss.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_loss.customContextMenuRequested.connect(self.table_right_menu)
        self.table_loss.setMinimumHeight(200)
        self.table_loss.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table_loss.setStyleSheet("""
            QTableWidget {background-color: white;border: 1px solid #D0E7FF;border-radius: 4px;gridline-color: #E0EFFF;}
            QHeaderView::section {background-color: #F5F9FF;border: 1px solid #D0E7FF;padding: 6px;font-weight: bold;}
            QTableWidget::item {padding: 4px;}
        """)

        btn_bar_h = QHBoxLayout()
        btn_bar_h.addStretch()
        self.btn_draw_bar = QPushButton("刷新横向柱状图与自动建议")
        self.btn_draw_bar.setStyleSheet("""
            QPushButton {padding: 6px 12px;border: 1px solid #D0E7FF;border-radius: 4px;background-color: white;}
            QPushButton:hover {background-color: #EBF4FF;}
        """)
        self.btn_draw_bar.clicked.connect(self.on_refresh_bar)
        btn_bar_h.addWidget(self.btn_draw_bar)

        panel_loss.add_layout(lay_loss_top)
        panel_loss.add_widget(self.table_loss)
        panel_loss.add_layout(btn_bar_h)
        self.btn_add_loss.clicked.connect(self.add_loss_row)
        self.btn_add_loss.clicked.connect(lambda: self.auto_open_panel("bottom"))
        self.btn_custom_loss.clicked.connect(self.add_new_loss)
        self.scroll_v.addWidget(panel_loss)

        panel_weak = CollapsiblePanel("【知识点强弱分析】", "weak", self)
        self.all_panels["weak"] = panel_weak
        top_h = QHBoxLayout()
        self.btn_import_excel = QPushButton("📂上传Excel知识点表格")
        self.btn_import_excel.setStyleSheet("""
            QPushButton {padding: 6px 12px;border: 1px solid #D0E7FF;border-radius: 4px;background-color: white;}
            QPushButton:hover {background-color: #EBF4FF;}
        """)
        self.btn_manual_add = QPushButton("➕手动新增知识点")
        self.btn_manual_add.setStyleSheet("""
            QPushButton {padding: 6px 12px;border: 1px solid #D0E7FF;border-radius: 4px;background-color: white;}
            QPushButton:hover {background-color: #EBF4FF;}
        """)
        top_h.addWidget(self.btn_import_excel)
        top_h.addWidget(self.btn_manual_add)

        self.scroll_know = QScrollArea()
        self.scroll_know.setWidgetResizable(True)
        self.scroll_know.setMinimumHeight(300)
        self.container_check = QWidget()
        self.layout_check = QVBoxLayout(self.container_check)
        self.layout_check.setAlignment(Qt.AlignTop)
        self.layout_check.setSpacing(4)
        self.scroll_know.setWidget(self.container_check)

        panel_weak.add_layout(top_h)
        panel_weak.add_widget(QLabel("☑勾选=优势知识点 | 不勾选=薄弱知识点"))
        panel_weak.add_widget(self.scroll_know)
        self.btn_import_excel.clicked.connect(self.read_excel_knowledge)
        self.btn_manual_add.clicked.connect(self.add_single_know)
        self.scroll_v.addWidget(panel_weak)

        panel_bottom = CollapsiblePanel("【课程总结填写】", "bottom", self)
        self.all_panels["bottom"] = panel_bottom
        lay_bot = QFormLayout()
        lay_bot.setLabelAlignment(Qt.AlignLeft)
        lay_bot.setSpacing(15)
        lay_bot.setHorizontalSpacing(26)

        h_prog_edit = QHBoxLayout()
        self.sp_total_lesson = NoWheelSpin()
        self.sp_total_lesson.setRange(0, 999)
        self.sp_done_lesson = NoWheelSpin()
        self.sp_done_lesson.setRange(0, 999)
        h_prog_edit.addWidget(QLabel("总课时："))
        h_prog_edit.addWidget(self.sp_total_lesson)
        h_prog_edit.addSpacing(15)
        h_prog_edit.addWidget(QLabel("已上课时："))
        h_prog_edit.addWidget(self.sp_done_lesson)
        lay_bot.addRow(QLabel("课程进度设置"), h_prog_edit)

        self.edit_level = NoWheelLineEdit()
        self.edit_core = NoWheelTextEdit()
        self.edit_suggest = NoWheelTextEdit()
        lay_bot.addRow(QLabel("能力评级："), self.edit_level)
        lay_bot.addRow(QLabel("核心问题："), self.edit_core)
        lay_bot.addRow(QLabel("教师建议："), self.edit_suggest)
        panel_bottom.add_layout(lay_bot)
        self.scroll_v.addWidget(panel_bottom)

        btn_export = QPushButton("🖼️ 导出预览为完整报告PNG")
        btn_export.setStyleSheet("background: #C8E6C9;color:black;padding:14px;font-size:18px;")
        btn_export.clicked.connect(self.save_preview_png)
        self.scroll_v.addWidget(btn_export)

        self.scroll_v.addStretch(1)
        self.scroll.setWidget(self.scroll_content)
        left_layout.addWidget(self.scroll)

        right_container = QWidget()
        right_layout = QVBoxLayout(right_container)
        right_layout.setContentsMargins(18, 30, 18, 18)

        self.scroll_preview = QScrollArea()
        self.scroll_preview.setWidgetResizable(True)
        self.scroll_preview.setStyleSheet("border:none;")
        self.scroll_preview.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.scroll_preview.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        self.preview_canvas = QWidget()
        self.preview_canvas.setStyleSheet("background:#ffffff;")
        self.preview_canvas.setFixedWidth(880)
        self.preview_canvas.setMinimumHeight(1500)

        self.report_progress = ReportProgressBar(self.preview_canvas)
        self.build_preview_layout()

        self.scroll_preview.setWidget(self.preview_canvas)
        right_layout.addWidget(self.scroll_preview)

        splitter_main.addWidget(left_container)
        splitter_main.addWidget(right_container)
        splitter_main.setSizes([500, 900])
        main_layout.addWidget(splitter_main)

        def refresh_preview_progress():
            total = self.sp_total_lesson.value()
            done = self.sp_done_lesson.value()
            course_name = self.in_course.text()
            per = done / total * 100 if total > 0 else 0.0
            self.report_progress.set_data(course_name, per)
            self.sync_preview()

        self.sp_total_lesson.valueChanged.connect(refresh_preview_progress)
        self.sp_done_lesson.valueChanged.connect(refresh_preview_progress)
        self.in_course.textChanged.connect(refresh_preview_progress)

    def clear_all_check(self):
        while self.layout_check.count() > 0:
            item = self.layout_check.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()
        self.know_check_list.clear()
        self.refresh_knowledge_height()

    def create_checkbox(self, name, is_checked=False):
        cb = KnowledgeCheckBox(str(name).strip())
        cb.setChecked(is_checked)
        self.layout_check.addWidget(cb)
        self.know_check_list.append(cb)
        self.refresh_knowledge_height()

    def add_single_know(self):
        name, ok = QInputDialog.getText(self, "新增知识点", "输入知识点名称：")
        if ok and name.strip():
            self.create_checkbox(name, False)

    def read_excel_knowledge(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel", "", "*.xlsx;*.xls")
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            self.clear_all_check()
            for row in range(1, ws.max_row + 1):
                kn_val = ws[f"A{row}"].value
                if kn_val is None or str(kn_val).strip() == "":
                    continue
                self.create_checkbox(kn_val, False)
        except Exception as e:
            QMessageBox.warning(self, "读取异常", f"表格读取失败：{str(e)}")

    def table_right_menu(self, pos):
        menu = QMenu()
        act_del = menu.addAction("删除当前行")
        act_edit = menu.addAction("修改选中内容")
        action = menu.exec_(self.table_loss.viewport().mapToGlobal(pos))
        row = self.table_loss.currentRow()
        col = self.table_loss.currentColumn()
        if action == act_del and row >= 0:
            self.table_loss.removeRow(row)
            self.cache_bar_text = ""
        elif action == act_edit and row >= 0 and col >= 0:
            old = self.table_loss.item(row, col).text() if self.table_loss.item(row, col) else ""
            new_txt, ok = QInputDialog.getText(self, "修改内容", "输入新内容：", text=old)
            if ok:
                self.table_loss.setItem(row, col, QTableWidgetItem(new_txt))
                self.cache_bar_text = ""

    def build_preview_layout(self):
        pv = self.preview_canvas
        self.title_main = QLabel("测试分析报告", pv)
        self.title_main.setGeometry(QRect(120, 12, 640, 72))
        self.title_main.setStyleSheet("font-size:48px; font-weight:bold; color:#2176D3; border:none;")
        self.title_main.setAlignment(Qt.AlignCenter)

        self.title_sub = QLabel("记录成长每一步·见证进步每一天", pv)
        self.title_sub.setGeometry(QRect(220, 84, 440, 28))
        self.title_sub.setStyleSheet("font-size:17px; color:#FF8C42; border:none;")
        self.title_sub.setAlignment(Qt.AlignCenter)

        self.card_stu = QGroupBox(pv)
        self.card_stu.setGeometry(QRect(28, 122, 352, 176))
        self.card_stu.setStyleSheet(f"border:3px solid {COLOR_BLUE_BORDER}; border-radius:16px; background:#F5FAFF;")
        self.text_stu = QLabel("", self.card_stu)
        self.text_stu.setGeometry(20, 10, 318, 148)
        self.text_stu.setStyleSheet("font-size:18px; color:#333333; border:none;")

        self.card_score = QGroupBox("综合得分", pv)
        self.card_score.setGeometry(QRect(392, 122, 460, 176))
        self.card_score.setStyleSheet(f"border:3px solid {COLOR_BLUE_BORDER}; border-radius:16px; background:#F5FAFF;")
        self.atext_score = QLabel("", self.card_score)
        self.atext_score.setGeometry(15, 25, 195, 142)
        self.atext_score.setStyleSheet("border:none;")
        self.btext_score = QLabel("", self.card_score)
        self.btext_score.setGeometry(233, 4, 195, 160)
        self.btext_score.setStyleSheet("border:none;")

        self.card_radar = QGroupBox("能力雷达图（五维能力）", pv)
        self.card_radar.setGeometry(QRect(28, 312, 362, 398))
        self.card_radar.setStyleSheet(
            f"border:3px solid {COLOR_YELLOW_BORDER}; border-radius:16px; background:#FFFDF5;")
        self.pre_can_radar = MplCanvasRadar(width=3.2, height=2.5)
        self.pre_can_radar.setParent(self.card_radar)
        self.pre_can_radar.setGeometry(QRect(5, 32, 350, 296))
        self.text_radar = QLabel("", self.card_radar)
        self.text_radar.setGeometry(17, 325, 322, 70)
        self.text_radar.setStyleSheet("font-size:14px; color:#333333; border:none;")

        self.card_pie = QGroupBox("知识点掌握分布", pv)
        self.card_pie.setGeometry(QRect(402, 312, 455, 398))
        self.card_pie.setStyleSheet(f"border:3px solid {COLOR_PURPLE_BORDER}; border-radius:16px; background:#FAF8FF;")
        self.pre_can_pie = MplCanvasPie(width=3.2, height=2.5)
        self.pre_can_pie.setParent(self.card_pie)
        self.pre_can_pie.setGeometry(QRect(8, 42, 268, 296))
        self.text_pie = QLabel("", self.card_pie)
        self.text_pie.setGeometry(QRect(285, 42, 163, 290))
        self.text_pie.setStyleSheet("font-size:18px; color:#333333; border:none;")

        self.card_bar = QGroupBox("失分原因", pv)
        self.card_bar.setGeometry(QRect(28, 722, 352, 408))
        self.card_bar.setStyleSheet(f"border:3px solid {COLOR_GREEN_BORDER}; border-radius:16px; background:#F5FFF5;")
        self.pre_can_bar = MplCanvasBar(width=3.2, height=2.2)
        self.pre_can_bar.setParent(self.card_bar)
        self.pre_can_bar.setGeometry(QRect(12, 30, 326, 266))
        self.text_bar = QLabel("", self.card_bar)
        self.text_bar.setGeometry(17, 290, 322, 114)
        self.text_bar.setStyleSheet("font-size:13px; color:#2176D3; border:none;")
        self.text_bar.setWordWrap(True)

        self.card_ana = QGroupBox("知识点强弱分析", pv)
        self.card_ana.setGeometry(QRect(392, 722, 460, 408))
        self.card_ana.setStyleSheet(f"border:3px solid {COLOR_ORANGE_BORDER}; border-radius:16px; background:#FFF9F0;")
        self.text_ana = QLabel("", self.card_ana)
        self.text_ana.setGeometry(18, 38, 426, 368)
        self.text_ana.setStyleSheet("font-size:18px; color:#333333; border:none;")
        self.text_ana.setWordWrap(True)

        self.card_bottom = QGroupBox(pv)
        self.card_bottom.setGeometry(QRect(28, 1154, 824, 228))
        self.card_bottom.setStyleSheet(
            f"border:3px solid {COLOR_BOTTOM_BORDER}; border-radius:16px; background:#F8F6FF;")
        self.report_progress.setParent(self.card_bottom)
        self.report_progress.setGeometry(QRect(12, 8, 798, 110))
        self.text_bottom = QLabel("", self.card_bottom)
        self.text_bottom.setGeometry(QRect(12, 72, 798, 138))
        self.text_bottom.setStyleSheet("font-size:14px; color:#333333; border:none;")
        self.text_bottom.setWordWrap(True)

        self.footer = QLabel("感谢您的信任与支持!让我们一起陪伴孩子快乐成长!", pv)
        self.footer.setGeometry(QRect(210, 1394, 460, 34))
        self.footer.setStyleSheet("font-size:15px; color:#2176D3; border:none;")
        self.footer.setAlignment(Qt.AlignCenter)

        #

    def on_refresh_radar(self):
        star_cnt, level_title, comment = self.get_radar_evaluation()
        self.edit_level.setText(level_title)

        loss_dict = {}
        for row in range(self.table_loss.rowCount()):
            reason = self.table_loss.item(row, 0).text() if self.table_loss.item(row, 0) else ""
            val = float(self.table_loss.item(row, 1).text()) if self.table_loss.item(row, 1) else 0
            if reason.strip():
                loss_dict[reason] = loss_dict.get(reason, 0) + val
        core_analysis, improve_analysis = self.generate_analysis(loss_dict)
        self.edit_core.setPlainText(core_analysis)
        final_suggest = f"{comment}\n【针对性改进建议】\n{improve_analysis}"
        self.edit_suggest.setPlainText(final_suggest)

        self.sync_preview()

    def on_refresh_bar(self):
        loss_dict = {}
        for row in range(self.table_loss.rowCount()):
            reason = self.table_loss.item(row, 0).text() if self.table_loss.item(row, 0) else ""
            val = float(self.table_loss.item(row, 1).text()) if self.table_loss.item(row, 1) else 0
            if reason.strip():
                loss_dict[reason] = loss_dict.get(reason, 0) + val
        core_analysis, improve_analysis = self.generate_analysis(loss_dict)
        self.edit_core.setPlainText(core_analysis)

        star_cnt, level_title, comment = self.get_radar_evaluation()
        self.edit_level.setText(level_title)
        final_suggest = f"{comment}\n【针对性改进建议】\n{improve_analysis}"
        self.edit_suggest.setPlainText(final_suggest)

        self.sync_preview()

    def sync_preview(self):
        try:
            #

            # 学生信息实时更新
            s_name = self.in_name.text() or "未填写"
            s_course = self.in_course.text() or "未填写"
            s_time = self.in_time.text() or "未填写"
            s_tea = self.in_teacher.text() or "未填写"

            txt = (
                f'<p style="margin:8px; line-height:20px; font-size:22px; color:#333333;">学生姓名：{s_name}</p>'
                f'<p style="margin:8px; line-height:20px; font-size:18px; color:#333333;">课程名称：{s_course}</p>'
                f'<p style="margin:8px; line-height:20px; font-size:18px; color:#333333;">测试时间：{s_time}</p>'
                f'<p style="margin:8px; line-height:20px; font-size:18px; color:#333333;">教师姓名：{s_tea}</p>'
            )
            self.text_stu.setText(txt)
            self.text_stu.repaint()  # 强制刷新UI

            # 综合得分实时更新
            full = self.sp_full.value()
            get = self.sp_get.value()
            obj_r = self.sp_obj_r.value()
            obj_w = self.sp_obj_w.value()

            self.prog_full_sum = 0.0
            self.prog_get_sum = 0.0
            prog_lines = []
            for idx, wid in enumerate(self.prog_item_list):
                lay = wid.layout()
                p_full = lay.itemAt(2).widget().value()
                p_get = lay.itemAt(4).widget().value()
                self.prog_full_sum += p_full
                self.prog_get_sum += p_get
                prog_lines.append(f"第{idx + 1}题   {int(p_get)}/{int(p_full)}")

            prog_str = "<br>".join(prog_lines) if prog_lines else "无题目"
            rate = f"{get / full * 100:.1f}%" if full > 0 else "0.00%"
            txt_a = (
                f'<p style="margin:0px;line-height:36px;font-size:40px;color:#222222;text-align:center;">{int(get)} / {int(full)}</p>'
                f'<p style="margin:0px;line-height:32px;font-size:18px;color:#2176D3;text-align:center;">得分率 {rate}</p>'
            )
            self.atext_score.setText(txt_a)
            self.atext_score.repaint()  # 强制刷新UI

            # 统一文字为"编程题"
            txt_b = (
                f'<p style="margin:0px;line-height:26px;font-size:20px;color:#333333;text-align:left;">客观题:'
                f'<span style="color:#4CAF50;">对 {obj_r}  '
                f'<span style="color:#FF7043;">错 {obj_w}</p>'
                f'<p style="margin:0px;line-height:26px;font-size:20px;color:#333333;text-align:left;">计算题:</p>'
                f'<p style="margin:0px;line-height:26px;font-size:20px;color:#4CAF50;text-align:left;">{prog_str}'
            )
            self.btext_score.setText(txt_b)
            self.btext_score.repaint()  # 强制刷新UI

            ax_r = self.pre_can_radar.axes
            ax_r.clear()
            labels = ["概念理解力", "计算能力", "模型应用", "审题能力", "知识体系"]
            scores = [self.radar_dict[name].value() for name in labels]
            ang = np.linspace(0, 2 * np.pi, 5, endpoint=False)
            ang_plot = list(ang) + [ang[0]]
            score_plot = scores + [scores[0]]
            ax_r.fill(ang_plot, score_plot, alpha=0.12, c="#2176D3")
            ax_r.plot(ang_plot, score_plot, color="#2176D3", linewidth=2)
            ax_r.fill(ang_plot, score_plot, alpha=0.25, c="#2176D3")
            ax_r.set_xticks(ang)
            ax_r.set_xticklabels(labels, fontsize=11, color="#333")
            ax_r.tick_params(axis='x', pad=6)
            ax_r.set_ylim(0, 100)
            ax_r.set_yticks([20, 40, 60, 80, 100])
            ax_r.set_yticklabels(["20", "40", "60", "80", "100"], fontsize=9)
            ax_r.grid(True, color="#cccccc")
            ax_r.spines["polar"].set_visible(False)
            ax_r.set_position([0.105, 0.12, 0.68, 0.76])
            self.pre_can_radar.draw()
            star_cnt, level_title, comment = self.get_radar_evaluation()
            star_str = "⭐" * star_cnt

            txt_star = (
                f'<p style="margin:0px;line-height:22px;font-size:20px;color:#333333;text-align:center;">能力水平：{star_str}</p>'
                f'<p style="margin:0px;line-height:22px;font-size:18px;color:#4CAF50;text-align:center;">{level_title}</p>'
            )
            self.text_radar.setText(txt_star)

            g = self.sp_k_good.value()
            m = self.sp_k_mid.value()
            b = self.sp_k_bad.value()
            allk = self.sp_k_all.value()
            data = [g, m, b]
            names = ["掌握较好", "中等掌握", "掌握较差"]
            colors = ["#A5D6A7", "#FFF59D", "#FFAB91"]
            ax_p = self.pre_can_pie.axes
            ax_p.clear()
            ax_p.axis("off")
            ax_p.set_position([0, 0.05, 0.90, 0.90])
            if sum(data) > 0:
                wedges, texts = ax_p.pie(data, labels=names, colors=colors, startangle=90,
                                         wedgeprops=dict(width=0.4), textprops={"fontsize": 9}, labeldistance=0.7)
            self.pre_can_pie.draw()
            if allk > 0 and sum(data) > 0:
                g_rate = round(g / allk * 100, 1)
                m_rate = round(m / allk * 100, 1)
                b_rate = round(b / allk * 100, 1)
                right_text = (f"✅掌握较好(L5)\n{g}个 | {g_rate}%\n理解扎实，运用自如\n\n"
                              f"🟡中等掌握\n{m}个 | {m_rate}%\n会做但不稳定\n\n"
                              f"🔴掌握较差\n{b}个 | {b_rate}%\n基础薄弱")
            else:
                right_text = "暂无数据，请填写知识点总数"
            self.text_pie.setText(right_text)

            ax_b = self.pre_can_bar.axes
            ax_b.clear()
            bar_names = []
            bar_vals = []
            for row in range(self.table_loss.rowCount()):
                reason = self.table_loss.item(row, 0).text() if self.table_loss.item(row, 0) else ""
                val = float(self.table_loss.item(row, 1).text()) if self.table_loss.item(row, 1) else 0
                if reason.strip():
                    bar_names.append(reason)
                    bar_vals.append(val)
            color_list = ["#FFCDD2", "#B2EBF2", "#BBDEFB", "#C8E6C9", "#FFF9C4", "#E1BEE7", "#B2DFDB"]
            if bar_names:
                use_color = color_list[:len(bar_names)]
                ax_b.barh(bar_names, bar_vals, color=use_color)
                ax_b.set_xlabel("占比(%)")
            ax_b.figure.tight_layout()
            self.pre_can_bar.draw()

            core_text = self.edit_core.toPlainText()
            visible_core = '\n'.join(core_text.split('\n')[:6])
            auto_text = (
                f'<p style="margin:0px;line-height:22px;font-size:16px;color:#2176D3;text-align:left;">【核心问题】</p>'
                f'<p style="margin:0px;line-height:22px;font-size:16px;color:#2176D3;text-align:left;">{visible_core}</p>'
            )
            self.text_bar.setText(auto_text)

            good_list = []
            bad_list = []
            for cb in self.know_check_list:
                name = cb.text()
                if cb.isChecked():
                    good_list.append(name)
                else:
                    bad_list.append(name)
            good_str = "\n".join(good_list) if good_list else "无"
            bad_str = "\n".join(bad_list) if bad_list else "无"
            preview_txt = f"✅优势知识点：\n{good_str}\n❌薄弱知识点：\n{bad_str}"
            self.text_ana.setText(preview_txt)

            lev = self.edit_level.text()
            current_suggest = self.edit_suggest.toPlainText()
            visible_sug = '\n'.join(current_suggest.split('\n')[:8])
            final_sug_html = visible_sug.replace("\n", "<br>")

            bottom_txt = (
                f'<p style="margin:0px;line-height:18px;font-size:18px;color:#482E6E;text-align:left;">能力评级：<span style="color:#2176D3">{lev}<br>'
                f'<span style="color:#482E6E">教师建议：</span>'
                f'<span style="color:#FF8C42;text-decoration:underline;text-underline-offset:2px;">{final_sug_html}</span></p>'
            )
            self.text_bottom.setText(bottom_txt)

            #

            # 强制刷新整个预览画布
            self.preview_canvas.repaint()

        except Exception as e:
            print("刷新异常：", e)
            import traceback
            traceback.print_exc()  # 打印详细错误信息，方便调试

    def load_image(self, label, path):
        if not path:
            label.clear()
            return
        try:
            img = QImage(path).convertToFormat(QImage.Format_ARGB32)
            pixmap = QPixmap.fromImage(img)
            scaled_pixmap = pixmap.scaled(label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            label.setPixmap(scaled_pixmap)
        except Exception as e:
            print(f"加载图片失败: {e}")
            label.clear()

    def create_prog_items(self):
        for w in self.prog_item_list:
            w.deleteLater()
        self.prog_item_list.clear()
        cnt = self.sp_prog_cnt.value()
        for i in range(cnt):
            self.add_single_prog_item(i + 1)
        self.scroll_content.updateGeometry()
        self.sync_preview()

    def add_single_prog_item(self, index):
        w = QWidget()
        hl = QHBoxLayout(w)
        hl.setContentsMargins(0, 0, 0, 0)
        hl.addWidget(QLabel(f"第{index}题"))
        hl.addWidget(QLabel("满分"))
        spf = NoWheelDoubleSpin()
        spf.setRange(0, 100)
        spf.valueChanged.connect(self.sync_preview)
        hl.addWidget(spf)
        hl.addWidget(QLabel("得分"))
        spg = NoWheelDoubleSpin()
        spg.setRange(0, 100)
        spg.valueChanged.connect(self.sync_preview)
        hl.addWidget(spg)

        btn_del = QPushButton("删除")
        btn_del.setStyleSheet("padding:4px 8px; background:#ff4444; color:white; border-radius:4px;")
        btn_del.clicked.connect(lambda: self.delete_prog_item(w))
        hl.addWidget(btn_del)

        self.lay_prog_wrap.addWidget(w)
        self.prog_item_list.append(w)

    def delete_prog_item(self, widget):
        if widget in self.prog_item_list:
            self.prog_item_list.remove(widget)
            widget.deleteLater()
            self.sync_preview()

    def add_loss_row(self):
        name = self.com_loss.currentText()
        per = self.sp_loss_per.value()
        row = self.table_loss.rowCount()
        self.table_loss.insertRow(row)
        self.table_loss.setItem(row, 0, QTableWidgetItem(name))
        self.table_loss.setItem(row, 1, QTableWidgetItem(str(per)))
        self.cache_bar_text = ""

    def add_new_loss(self):
        txt, ok = QInputDialog.getText(self, "自定义失分原因", "输入名称：")
        if ok and txt.strip():
            row = self.table_loss.rowCount()
            self.table_loss.insertRow(row)
            self.table_loss.setItem(row, 0, QTableWidgetItem(txt.strip()))
            self.table_loss.setItem(row, 1, QTableWidgetItem("0"))
            self.cache_bar_text = ""

    def save_preview_png(self):
        path, _ = QFileDialog.getSaveFileName(self, "保存报告", "测试报告.png", "PNG图片 (*.png)")
        if not path:
            return
        try:
            self.preview_canvas.setMinimumSize(self.preview_canvas.sizeHint())
            QApplication.processEvents()
            pix = QPixmap(self.preview_canvas.size())
            self.preview_canvas.render(pix)
            pix.save(path)
            QMessageBox.information(self, "成功", f"报告已保存：\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "失败", f"保存出错：{str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = MainWin()
    win.show()
    sys.exit(app.exec_())
